import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, PatternFill
from datetime import datetime

ARQUIVO_EXCEL = "data/dados_climaticos.xlsx"


def salvar_dados(cidade, temperatura, umidade, ceu):

    if not os.path.exists(ARQUIVO_EXCEL):
        arquivo = Workbook()
        planilha = arquivo.active
        planilha.title = "Clima"

        cabecalhos = ["Data e Hora", "Cidade", "Temperatura (°C)", "Umidade (%)", "Condição do Céu"]
        for col, cabecalho in enumerate(cabecalhos, start=1):
            cell = planilha.cell(row=1, column=col)
            cell.value = cabecalho
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            cell.fill = PatternFill(start_color="FFD700", fill_type="solid")
    else:
        arquivo = load_workbook(ARQUIVO_EXCEL)
        planilha = arquivo["Clima"]

    linha = planilha.max_row + 1
    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    valores = [data_hora, cidade, temperatura, umidade, ceu]

    for col, valor in enumerate(valores, start=1):
        cell = planilha.cell(row=linha, column=col)
        cell.value = valor
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    arquivo.save(ARQUIVO_EXCEL)
